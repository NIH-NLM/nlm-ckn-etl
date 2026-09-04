import json
import sys
import tempfile
import unittest
from pathlib import Path

# Both directories: the module itself was retired to src/_deprecated/, but it
# still imports LoaderUtilities from src/.
_PYTHON_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(_PYTHON_ROOT / "src"))
sys.path.insert(0, str(_PYTHON_ROOT / "src" / "_deprecated"))

from openpyxl import load_workbook

from UberonHuBMAPAuditor import (
    COLUMNS_SHEET,
    COLUMN_SOURCES,
    LABEL_CONFLICTS_SHEET,
    NEW_PART_OF_SHEET,
    SUMMARY_SHEET,
    UNKNOWN_TERMS_SHEET,
    audit,
    find_label_conflicts,
    find_new_part_of_edges,
    columns_of,
    find_unknown_terms,
    index_hubmap,
    index_uberon,
)

PURL = "http://purl.obolibrary.org/obo"
PART_OF = f"{PURL}/BFO_0000050"
RDFS_LABEL = "http://www.w3.org/2000/01/rdf-schema#label"
RDF_LABEL = "http://www.w3.org/1999/02/22-rdf-syntax-ns#label"

# UBERON: kidney part_of urinary system, and a nephron with no HuBMAP counterpart
UBERON_TUPLES = [
    [f"{PURL}/UBERON_0002113", RDFS_LABEL, "kidney"],
    [f"{PURL}/UBERON_0002113", PART_OF, f"{PURL}/UBERON_0013702"],
    [f"{PURL}/UBERON_0013702", RDFS_LABEL, "body proper"],
    [f"{PURL}/UBERON_0001285", RDFS_LABEL, "nephron"],
    [f"{PURL}/UBERON_0002015", RDFS_LABEL, "kidney capsule"],
]

# HuBMAP: agrees about the kidney edge, relabels the kidney capsule, adds a
# part_of edge UBERON does not assert, and mentions a term UBERON does not have
HUBMAP_TUPLES = [
    [f"{PURL}/UBERON_0002113", PART_OF, f"{PURL}/UBERON_0013702"],
    [f"{PURL}/UBERON_0002113", RDF_LABEL, "kidney"],
    [f"{PURL}/UBERON_0002015", PART_OF, f"{PURL}/UBERON_0002113"],
    [f"{PURL}/UBERON_0002015", RDF_LABEL, "renal capsule"],
    [f"{PURL}/UBERON_9999999", PART_OF, f"{PURL}/UBERON_0002113"],
    [f"{PURL}/UBERON_9999999", RDF_LABEL, "made up structure"],
]


class UberonHuBMAPAuditorTestCase(unittest.TestCase):
    """Tests for the UBERON and HuBMAP audit."""

    def setUp(self):
        self.uberon = index_uberon(UBERON_TUPLES)
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.hubmap_path = Path(self.tmp_dir.name) / "hubmap-kidney-v1.8.json"
        with open(self.hubmap_path, "w") as fp:
            json.dump({"tuples": HUBMAP_TUPLES}, fp)
        self.hubmap = index_hubmap([self.hubmap_path])

    def tearDown(self):
        self.tmp_dir.cleanup()

    def test_index_uberon(self):
        self.assertEqual(self.uberon["labels"]["UBERON_0002113"], {"kidney"})
        self.assertEqual(self.uberon["part_of"]["UBERON_0002113"], {"UBERON_0013702"})
        self.assertIn("UBERON_0001285", self.uberon["terms"])

    def test_index_hubmap_carries_source(self):
        self.assertEqual(
            self.hubmap["labels"]["UBERON_0002015"]["renal capsule"],
            {"hubmap-kidney-v1.8"},
        )
        self.assertEqual(
            self.hubmap["part_of"][("UBERON_0002015", "UBERON_0002113")],
            {"hubmap-kidney-v1.8"},
        )

    def test_find_label_conflicts(self):
        conflicts = find_label_conflicts(self.uberon, self.hubmap)
        self.assertEqual(len(conflicts), 1)
        row = conflicts.iloc[0]
        self.assertEqual(row["uberon_term"], "UBERON:0002015")
        self.assertEqual(row["uberon_label"], "kidney capsule")
        self.assertEqual(row["hubmap_label"], "renal capsule")
        self.assertEqual(row["audit_differs_by"], "substantive")
        self.assertEqual(row["hubmap_source"], "hubmap-kidney-v1.8")

    def test_find_label_conflicts_ignores_terms_uberon_does_not_label(self):
        conflicts = find_label_conflicts(self.uberon, self.hubmap)
        self.assertNotIn("UBERON:9999999", set(conflicts["uberon_term"]))

    def test_find_label_conflicts_flags_case_only_difference(self):
        hubmap = index_hubmap([self.hubmap_path])
        hubmap["labels"]["UBERON_0002113"] = {"Kidney": {"hubmap-kidney-v1.8"}}
        conflicts = find_label_conflicts(self.uberon, hubmap)
        row = conflicts[conflicts["uberon_term"] == "UBERON:0002113"].iloc[0]
        self.assertEqual(row["audit_differs_by"], "case-or-whitespace")

    def test_find_new_part_of_edges(self):
        edges = find_new_part_of_edges(self.uberon, self.hubmap)
        pairs = set(zip(edges["hubmap_subject_term"], edges["hubmap_object_term"]))
        # The kidney edge is asserted by UBERON too, so it is not reported
        self.assertNotIn(("UBERON:0002113", "UBERON:0013702"), pairs)
        self.assertEqual(
            pairs,
            {
                ("UBERON:0002015", "UBERON:0002113"),
                ("UBERON:9999999", "UBERON:0002113"),
            },
        )

    def test_find_new_part_of_edges_carries_uberon_parents(self):
        edges = find_new_part_of_edges(self.uberon, self.hubmap)
        row = edges[edges["hubmap_subject_term"] == "UBERON:0002015"].iloc[0]
        self.assertEqual(row["subject_label"], "kidney capsule")
        self.assertEqual(row["subject_label_source"], "UBERON")
        self.assertEqual(row["object_label"], "kidney")
        self.assertEqual(row["uberon_parents"], "")

    def test_find_new_part_of_edges_names_the_source_of_each_label(self):
        edges = find_new_part_of_edges(self.uberon, self.hubmap)
        row = edges[edges["hubmap_subject_term"] == "UBERON:9999999"].iloc[0]
        self.assertEqual(row["subject_label"], "made up structure")
        self.assertEqual(row["subject_label_source"], "HuBMAP")
        # The object is labelled by UBERON, so its label comes from there
        self.assertEqual(row["object_label"], "kidney")
        self.assertEqual(row["object_label_source"], "UBERON")

    def test_find_unknown_terms(self):
        unknown = find_unknown_terms(self.uberon, self.hubmap, deprecated_terms=[])
        self.assertEqual(len(unknown), 1)
        row = unknown.iloc[0]
        self.assertEqual(row["hubmap_term"], "UBERON:9999999")
        self.assertEqual(row["audit_reason"], "absent-from-uberon")

    def test_find_unknown_terms_flags_deprecated(self):
        unknown = find_unknown_terms(
            self.uberon, self.hubmap, deprecated_terms=["UBERON_0002015"]
        )
        reasons = dict(zip(unknown["hubmap_term"], unknown["audit_reason"]))
        self.assertEqual(reasons["UBERON:0002015"], "deprecated")

    def test_audit_writes_workbook(self):
        uberon_path = Path(self.tmp_dir.name) / "uberon-tuples.json"
        with open(uberon_path, "w") as fp:
            json.dump({"tuples": UBERON_TUPLES}, fp)
        output_path = Path(self.tmp_dir.name) / "uberon-hubmap-review.xlsx"

        sheets = audit(uberon_path, [self.hubmap_path], output_path)

        self.assertEqual(
            set(sheets),
            {LABEL_CONFLICTS_SHEET, NEW_PART_OF_SHEET, UNKNOWN_TERMS_SHEET},
        )
        workbook = load_workbook(output_path)
        self.assertEqual(
            workbook.sheetnames,
            [
                SUMMARY_SHEET,
                COLUMNS_SHEET,
                LABEL_CONFLICTS_SHEET,
                NEW_PART_OF_SHEET,
                UNKNOWN_TERMS_SHEET,
            ],
        )
        worksheet = workbook[LABEL_CONFLICTS_SHEET]
        header = [cell.value for cell in worksheet[1]]
        self.assertEqual(header, columns_of(LABEL_CONFLICTS_SHEET))
        self.assertEqual(header[-1], "Decision")
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(len(worksheet.data_validations.dataValidation), 1)

    def test_audit_writes_column_sources(self):
        uberon_path = Path(self.tmp_dir.name) / "uberon-tuples.json"
        with open(uberon_path, "w") as fp:
            json.dump({"tuples": UBERON_TUPLES}, fp)
        output_path = Path(self.tmp_dir.name) / "uberon-hubmap-review.xlsx"

        sheets = audit(uberon_path, [self.hubmap_path], output_path)

        # Every column of every detail sheet is documented, in sheet order
        for name, rows in sheets.items():
            self.assertEqual(list(rows.columns), columns_of(name))

        worksheet = load_workbook(output_path)[COLUMNS_SHEET]
        header = [cell.value for cell in worksheet[1]]
        self.assertEqual(header, ["Sheet", "Column", "Source", "Holds"])
        documented = {
            (row[0].value, row[1].value) for row in worksheet.iter_rows(min_row=2)
        }
        self.assertEqual(
            documented,
            {
                (name, column)
                for name, columns in COLUMN_SOURCES.items()
                for column, _, _ in columns
            },
        )


if __name__ == "__main__":
    unittest.main()
